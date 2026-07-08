import openpyxl, sys, json
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('job_api_catalog.xlsx', data_only=True)
ws = wb['API Catalog']
headers = [c.value for c in ws[1]]
idx = {h:i for i,h in enumerate(headers)}

def is_unknown(v):
    return isinstance(v,str) and v.strip().lower()=='unknown'
def is_blank(v):
    return v is None or (isinstance(v,str) and v.strip()=='')

# Columns we will NOT research - correct-by-design per Schema Definitions sheet
SKIP_COLS = {'Unique Job Identifier', 'Notes'}

CONTEXT_COLS = ['API/ATS Name','Region','Type','Access Model','Cost Details','Auth Required',
                'Docs Available','Notes','Status']

records = []
for r in range(2, ws.max_row+1):
    vals = [ws.cell(row=r, column=c+1).value for c in range(len(headers))]
    if all(v is None for v in vals):
        continue
    row = dict(zip(headers, vals))
    missing = [h for h in headers if h not in SKIP_COLS and (is_unknown(row[h]) or is_blank(row[h]) and h != 'Notes')]
    if not missing:
        continue
    records.append({'excel_row': r, 'row': row, 'missing_columns': missing})

print('rows needing real research:', len(records))
from collections import Counter
colcount = Counter()
for rec in records:
    for c in rec['missing_columns']:
        colcount[c]+=1
print(colcount)

# group by provider name
from collections import OrderedDict
providers = OrderedDict()
for rec in records:
    name = rec['row']['API/ATS Name']
    providers.setdefault(name, []).append(rec)
print('unique providers needing research:', len(providers))

with open('scratch_research_targets.json','w',encoding='utf-8') as f:
    json.dump(providers, f, indent=1, default=str)
