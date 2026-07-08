import openpyxl, sys, json
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('job_api_catalog.xlsx', data_only=False)
ws = wb['API Catalog']
headers = [c.value for c in ws[1]]
idx = {h:i for i,h in enumerate(headers)}

def is_unknown(v):
    return isinstance(v,str) and v.strip().lower()=='unknown'

NA_ELIGIBLE_COLS = ['Pagination','Pagination Type','API Limits','Incremental Sync',
                     'Date-based Sync','Limitations / Rate Limits','Additional Supported Regions']

def no_api_status(status):
    if not status: return False
    s = status.lower()
    return ('no public api' in s) or ('no public read api' in s)

na_fill_log = []
for r in range(2, ws.max_row+1):
    status = ws.cell(row=r, column=idx['Status']+1).value
    name = ws.cell(row=r, column=idx['API/ATS Name']+1).value
    if not name:
        continue
    if not no_api_status(status):
        continue
    for col in NA_ELIGIBLE_COLS:
        cidx = idx[col]+1
        cell = ws.cell(row=r, column=cidx)
        if is_unknown(cell.value):
            na_fill_log.append({'row':r,'name':name,'col':col,'old':cell.value,'status':status})
            cell.value = 'N/A - no public read API'

print('Total N/A fills applied:', len(na_fill_log))
from collections import Counter
print(Counter(l['col'] for l in na_fill_log))

wb.save('job_api_catalog.xlsx')
with open('scratch_na_fill_log.json','w',encoding='utf-8') as f:
    json.dump(na_fill_log, f, indent=1, default=str)
